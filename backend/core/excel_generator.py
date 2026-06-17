"""
Excel report generator using openpyxl.

Sheets:
  - Summary    — high-level KPIs
  - Monitors   — per-monitor uptime data
  - Incidents  — incident list

Styling:
  - Dark header row (#1e293b) with bold white text
  - Alternating row fill (#f1f5f9 / white)
  - Auto column widths
"""
import io
from datetime import datetime
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter

# ── Brand colours (openpyxl uses ARGB hex strings) ────────────────────────────
FILL_HEADER = PatternFill("solid", fgColor="1E293B")  # #1e293b
FILL_ALT = PatternFill("solid", fgColor="F1F5F9")     # #f1f5f9
FILL_WHITE = PatternFill("solid", fgColor="FFFFFF")
FILL_GREEN = PatternFill("solid", fgColor="D1FAE5")   # light green tint
FILL_RED = PatternFill("solid", fgColor="FEE2E2")     # light red tint

FONT_HEADER = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
FONT_BODY = Font(name="Calibri", size=9, color="1E293B")
FONT_BODY_BOLD = Font(name="Calibri", bold=True, size=9, color="1E293B")

ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=False)
ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=False)

THIN_SIDE = Side(style="thin", color="E2E8F0")
THIN_BORDER = Border(
    left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE
)


def _apply_header(ws, headers: list[str], row: int = 1) -> None:
    """Write a styled header row."""
    for col, label in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=label)
        cell.fill = FILL_HEADER
        cell.font = FONT_HEADER
        cell.alignment = ALIGN_CENTER
        cell.border = THIN_BORDER


def _apply_data_row(ws, values: list, row: int, alt: bool = False) -> None:
    """Write a data row with alternating fill."""
    fill = FILL_ALT if alt else FILL_WHITE
    for col, val in enumerate(values, start=1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.fill = fill
        cell.font = FONT_BODY
        cell.alignment = ALIGN_LEFT
        cell.border = THIN_BORDER


def _auto_width(ws) -> None:
    """Fit column widths to content (with a sensible cap)."""
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 50)


def _freeze_header(ws) -> None:
    ws.freeze_panes = "A2"


class ExcelGenerator:
    """Generates branded Excel reports as bytes."""

    def generate_uptime_report(
        self,
        org: Any,
        monitors_data: list[dict],
        period: str,
    ) -> bytes:
        wb = Workbook()

        # ── Summary sheet ──────────────────────────────────────────────────
        ws_sum = wb.active
        ws_sum.title = "Summary"
        org_name = getattr(org, "name", str(org)) if org else "All Monitors"
        total = len(monitors_data)
        sla_ok = sum(1 for m in monitors_data if m.get("sla_achieved"))
        avg_up = (
            sum(m.get("uptime_pct", 0) for m in monitors_data) / total if total else 0
        )
        avg_rt = (
            sum(m.get("avg_response_time", 0) for m in monitors_data) / total if total else 0
        )

        _apply_header(ws_sum, ["Metric", "Value"])
        kpis = [
            ("Organisation", org_name),
            ("Report Period", period),
            ("Generated At", datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC")),
            ("Total Monitors", total),
            ("Monitors Meeting SLA (≥99.9%)", sla_ok),
            ("Average Uptime (%)", f"{avg_up:.2f}"),
            ("Average Response Time (ms)", f"{avg_rt:.0f}"),
        ]
        for i, (k, v) in enumerate(kpis, start=2):
            _apply_data_row(ws_sum, [k, v], row=i, alt=(i % 2 == 0))
        _auto_width(ws_sum)

        # ── Monitors sheet ─────────────────────────────────────────────────
        ws_mon = wb.create_sheet("Monitors")
        headers = [
            "Monitor Name",
            "URL",
            "Uptime %",
            "Downtime %",
            "Avg Response Time (ms)",
            "Total Checks",
            "Down Checks",
            "SLA Achieved",
        ]
        _apply_header(ws_mon, headers)
        _freeze_header(ws_mon)
        for i, m in enumerate(monitors_data, start=2):
            sla_val = "YES" if m.get("sla_achieved") else "NO"
            row_data = [
                m.get("name", ""),
                m.get("url", ""),
                round(m.get("uptime_pct", 0), 2),
                round(m.get("downtime_pct", 0), 2),
                round(m.get("avg_response_time", 0), 0),
                m.get("total_checks", 0),
                m.get("down_checks", 0),
                sla_val,
            ]
            _apply_data_row(ws_mon, row_data, row=i, alt=(i % 2 == 0))
            # Colour the SLA cell
            sla_cell = ws_mon.cell(row=i, column=8)
            sla_cell.fill = FILL_GREEN if m.get("sla_achieved") else FILL_RED
            sla_cell.font = FONT_BODY_BOLD
            sla_cell.alignment = ALIGN_CENTER
        _auto_width(ws_mon)

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    def generate_incident_report(
        self,
        org: Any,
        incidents_data: dict,
        period: str,
    ) -> bytes:
        wb = Workbook()
        org_name = getattr(org, "name", str(org)) if org else "All Monitors"

        # ── Summary sheet ──────────────────────────────────────────────────
        ws_sum = wb.active
        ws_sum.title = "Summary"
        _apply_header(ws_sum, ["Metric", "Value"])
        kpis = [
            ("Organisation", org_name),
            ("Report Period", period),
            ("Generated At", datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC")),
            ("Total Incidents", incidents_data.get("total_incidents", 0)),
            ("Resolved", incidents_data.get("resolved", 0)),
            ("Ongoing", incidents_data.get("ongoing", 0)),
            (
                "Avg Resolution Time (mins)",
                f"{incidents_data.get('avg_resolution_mins', 0):.1f}",
            ),
        ]
        for i, (k, v) in enumerate(kpis, start=2):
            _apply_data_row(ws_sum, [k, v], row=i, alt=(i % 2 == 0))
        _auto_width(ws_sum)

        # ── Incidents sheet ────────────────────────────────────────────────
        ws_inc = wb.create_sheet("Incidents")
        headers = [
            "ID",
            "Monitor Name",
            "Started At",
            "Resolved At",
            "Duration (mins)",
            "Status",
            "Error Message",
        ]
        _apply_header(ws_inc, headers)
        _freeze_header(ws_inc)
        incidents = incidents_data.get("incidents", [])
        for i, inc in enumerate(incidents, start=2):
            row_data = [
                inc.get("id", ""),
                inc.get("monitor_name", ""),
                str(inc.get("started_at", ""))[:19],
                str(inc.get("resolved_at", "") or "Ongoing")[:19],
                inc.get("duration_mins", ""),
                inc.get("status", ""),
                inc.get("error", "") or "",
            ]
            _apply_data_row(ws_inc, row_data, row=i, alt=(i % 2 == 0))
            status_cell = ws_inc.cell(row=i, column=6)
            if inc.get("status") == "resolved":
                status_cell.fill = FILL_GREEN
            else:
                status_cell.fill = FILL_RED
            status_cell.font = FONT_BODY_BOLD
        _auto_width(ws_inc)

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    def generate_summary_report(
        self,
        org: Any,
        summary_data: dict,
        period: str,
    ) -> bytes:
        wb = Workbook()
        org_name = getattr(org, "name", str(org)) if org else "All Monitors"

        ws_sum = wb.active
        ws_sum.title = "Executive Summary"
        _apply_header(ws_sum, ["KPI", "Value"])
        kpis = [
            ("Organisation", summary_data.get("org_name", org_name)),
            ("Report Period", summary_data.get("period", period)),
            ("Generated At", datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC")),
            ("Total Monitors", summary_data.get("total_monitors", 0)),
            ("Average Uptime (%)", f"{summary_data.get('avg_uptime', 0):.2f}"),
            (
                "SLA Compliance (%)",
                f"{summary_data.get('sla_compliance_pct', 0):.1f}",
            ),
            ("Total Incidents", summary_data.get("total_incidents", 0)),
            (
                "Avg Response Time (ms)",
                f"{summary_data.get('avg_response_time', 0):.0f}",
            ),
        ]
        for i, (k, v) in enumerate(kpis, start=2):
            _apply_data_row(ws_sum, [k, v], row=i, alt=(i % 2 == 0))
        _auto_width(ws_sum)

        # Performers sheet
        ws_perf = wb.create_sheet("Performers")
        _apply_header(ws_perf, ["Category", "Monitor Name", "Uptime %"])
        row = 2
        best = summary_data.get("best_monitor")
        worst = summary_data.get("worst_monitor")
        if best:
            _apply_data_row(
                ws_perf,
                ["Best", best.get("name", ""), f"{best.get('uptime', 0):.2f}"],
                row=row,
            )
            ws_perf.cell(row=row, column=1).fill = FILL_GREEN
            row += 1
        if worst:
            _apply_data_row(
                ws_perf,
                ["Worst", worst.get("name", ""), f"{worst.get('uptime', 0):.2f}"],
                row=row,
            )
            ws_perf.cell(row=row, column=1).fill = FILL_RED
        _auto_width(ws_perf)

        # Recommendations sheet
        recs = summary_data.get("recommendations", [])
        if recs:
            ws_rec = wb.create_sheet("Recommendations")
            _apply_header(ws_rec, ["#", "Recommendation"])
            for i, rec in enumerate(recs, start=2):
                _apply_data_row(ws_rec, [i - 1, rec], row=i, alt=(i % 2 == 0))
            _auto_width(ws_rec)

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()
